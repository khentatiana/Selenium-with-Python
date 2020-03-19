import openpyxl

path = "/Users/tatiana/Desktop/PROJECTS/Selenium_Python_on_Youtube/For_testing_copy.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active

for r in range(1, 4): # will create 3 rows = 4 - 1
    for c in range(1, 6): # will create 5 columns = 6 - 1
        sheet.cell(row=1, column=1).value="Client_ID"
        sheet.cell(row=1, column=2).value = "Client_Name"
        sheet.cell(row=1, column=3).value = "Client_Last Name"
        sheet.cell(row=2, column=1).value = "1001"
        sheet.cell(row=2, column=2).value = "Grigori"
        sheet.cell(row=2, column=3).value = "Barbachov"

workbook.save(path)