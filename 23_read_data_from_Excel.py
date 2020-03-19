import openpyxl

#specify path to excel file
path = "/Users/tatiana/Desktop/PROJECTS/Selenium_Python_on_Youtube/For_testing.xlsx"
#load workbook(create workbook object)
workbook = openpyxl.load_workbook(path)
#read active sheets from excel file if only one active (create sheet object)
sheet = workbook.active
#read  sheets from excel file by name
#sheet = workbook.get_sheet_by_name("SheetName")

#number of rows and columnc in excel files
rows = sheet.max_row #6
columns = sheet.max_column #3

print("Number of rows:", rows)
print("Number of columns:", columns)

print("Content of excel file as following:")
#to show content of excel file
for r in range(1, rows + 1):
    for c in range(1, columns +1):
        print(sheet.cell(row=r, column=c).value, end=" ")
    print()

