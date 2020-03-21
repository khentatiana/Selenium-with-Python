import openpyxl

path = "/Users/tatiana/Desktop/PROJECTS/Selenium_Python_on_Youtube/For_testing_copy2.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active

for r in range(1, 4): # will create 3 rows = 4 - 1
    for c in range(1, 6): # will create 5 columns = 6 - 1
        sheet.cell(row=1, column=1).value="UserName"
        sheet.cell(row=1, column=2).value = "Password"
        sheet.cell(row=1, column=3).value = "Results"
        sheet.cell(row=2, column=1).value = "mercury"
        sheet.cell(row=2, column=2).value = "mercury"
        sheet.cell(row=2, column=3).value = ""
        sheet.cell(row=3, column=1).value = "mercury"
        sheet.cell(row=3, column=2).value = "asdasdA"
        sheet.cell(row=3, column=3).value = ""
        sheet.cell(row=4, column=1).value = "mno"
        sheet.cell(row=4, column=2).value = "asdasdA"
        sheet.cell(row=4, column=3).value = ""
        sheet.cell(row=5, column=1).value = "mno"
        sheet.cell(row=5, column=2).value = "asdasdA"
        sheet.cell(row=5, column=3).value = ""
        sheet.cell(row=6, column=1).value = "mercury"
        sheet.cell(row=6, column=2).value = "mercury"
        sheet.cell(row=6, column=3).value = ""

workbook.save(path)